import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from io import BytesIO
from datetime import datetime
from typing import BinaryIO
from models.property import Property
from models.document import UnderwritingAnalysis

class ExcelGenerator:
    def __init__(self):
        self.workbook = None
        
    def generate_underwriting_excel(
        self, 
        property: Property, 
        analysis: UnderwritingAnalysis
    ) -> BytesIO:
        """Generate comprehensive underwriting Excel file."""
        
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create worksheets
        self._create_executive_summary(property, analysis)
        self._create_cash_flow_analysis(property, analysis)
        self._create_10_year_projection(property, analysis)
        self._create_ratios_metrics(property, analysis)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _create_executive_summary(self, property: Property, analysis: UnderwritingAnalysis):
        """Create executive summary worksheet."""
        ws = self.workbook.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "INVESTMENT PROPERTY UNDERWRITING ANALYSIS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        row = 3
        
        # Property Information
        ws[f'A{row}'] = "Property Information"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        property_info = [
            ("Address:", property.address),
            ("City:", property.city),
            ("State:", property.state),
            ("ZIP Code:", property.zip_code),
            ("Units:", property.units),
            ("Property Type:", property.property_type.value),
            ("Year Built:", property.year_built or 'N/A'),
            ("Square Footage:", property.square_footage or 'N/A'),
        ]
        
        for label, value in property_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Purchase Information
        ws[f'A{row}'] = "Purchase Information"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        purchase_info = [
            ("Purchase Price:", f"${analysis.purchase_price:,.0f}"),
            ("Down Payment:", f"${analysis.down_payment:,.0f}"),
            ("Loan Amount:", f"${analysis.loan_amount:,.0f}"),
            ("Interest Rate:", f"{analysis.interest_rate:.2f}%"),
            ("Loan Term:", f"{analysis.loan_term} years"),
        ]
        
        for label, value in purchase_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Key Metrics
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        metrics = [
            ("Cap Rate:", f"{analysis.calculations.cap_rate:.2f}%"),
            ("Cash-on-Cash Return:", f"{analysis.calculations.cash_on_cash_return:.2f}%"),
            ("Debt Service Coverage:", f"{analysis.calculations.debt_service_coverage:.2f}"),
            ("Monthly Cash Flow:", f"${analysis.calculations.monthly_cash_flow:,.0f}"),
            ("Annual Cash Flow:", f"${analysis.calculations.annual_cash_flow:,.0f}"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            # Highlight key metrics
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
    
    def _create_cash_flow_analysis(self, property: Property, analysis: UnderwritingAnalysis):
        """Create cash flow analysis worksheet."""
        ws = self.workbook.create_sheet("Cash Flow Analysis")
        
        # Title
        ws['A1'] = "CASH FLOW ANALYSIS"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Income section
        ws[f'A{row}'] = "Income"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        income_items = [
            ("Gross Rental Income (Monthly):", f"${analysis.monthly_rent:,.0f}"),
            ("Gross Rental Income (Annual):", f"${analysis.calculations.gross_rental_income:,.0f}"),
            (f"Less: Vacancy ({analysis.vacancy}%):", f"-${analysis.calculations.gross_rental_income * analysis.vacancy / 100:,.0f}"),
            ("Effective Rental Income:", f"${analysis.calculations.gross_rental_income * (1 - analysis.vacancy / 100):,.0f}"),
        ]
        
        for label, value in income_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Operating Expenses
        ws[f'A{row}'] = "Operating Expenses"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        expenses = [
            ("Management:", f"${analysis.operating_expenses.management:,.0f}"),
            ("Maintenance & Repairs:", f"${analysis.operating_expenses.maintenance:,.0f}"),
            ("Insurance:", f"${analysis.operating_expenses.insurance:,.0f}"),
            ("Property Taxes:", f"${analysis.operating_expenses.taxes:,.0f}"),
            ("Utilities:", f"${analysis.operating_expenses.utilities:,.0f}"),
            ("Other Expenses:", f"${analysis.operating_expenses.other:,.0f}"),
        ]
        
        total_expenses = sum([
            analysis.operating_expenses.management,
            analysis.operating_expenses.maintenance,
            analysis.operating_expenses.insurance,
            analysis.operating_expenses.taxes,
            analysis.operating_expenses.utilities,
            analysis.operating_expenses.other,
        ])
        
        for label, value in expenses:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        ws[f'A{row}'] = "Total Operating Expenses:"
        ws[f'B{row}'] = f"${total_expenses:,.0f}"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # NOI and Cash Flow
        ws[f'A{row}'] = "Net Operating Income (NOI):"
        ws[f'B{row}'] = f"${analysis.calculations.net_operating_income:,.0f}"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Debt Service
        ws[f'A{row}'] = "Debt Service"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        debt_items = [
            ("Monthly Debt Service:", f"${analysis.calculations.monthly_debt_service:,.0f}"),
            ("Annual Debt Service:", f"${analysis.calculations.monthly_debt_service * 12:,.0f}"),
        ]
        
        for label, value in debt_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Final Cash Flow
        ws[f'A{row}'] = "Cash Flow"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        cash_flow_items = [
            ("Before-Tax Cash Flow (Annual):", f"${analysis.calculations.annual_cash_flow:,.0f}"),
            ("Before-Tax Cash Flow (Monthly):", f"${analysis.calculations.monthly_cash_flow:,.0f}"),
        ]
        
        for label, value in cash_flow_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            # Highlight cash flow
            ws[f'B{row}'].font = Font(bold=True)
            if "Annual" in label:
                # Color code positive/negative cash flow
                if analysis.calculations.annual_cash_flow > 0:
                    ws[f'B{row}'].font = Font(bold=True, color="00800000")  # Green
                else:
                    ws[f'B{row}'].font = Font(bold=True, color="00FF0000")  # Red
            row += 1
    
    def _create_10_year_projection(self, property: Property, analysis: UnderwritingAnalysis):
        """Create 10-year projection worksheet."""
        ws = self.workbook.create_sheet("10-Year Projection")
        
        # Title
        ws['A1'] = "10-YEAR CASH FLOW PROJECTION"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ["Year", "Rental Income", "Operating Expenses", "NOI", "Debt Service", "Cash Flow", "Cumulative CF"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Projection calculations
        cumulative_cf = 0
        rent_growth = 0.03  # 3% annual rent growth
        expense_growth = 0.02  # 2% annual expense growth
        
        base_expenses = sum([
            analysis.operating_expenses.management,
            analysis.operating_expenses.maintenance,
            analysis.operating_expenses.insurance,
            analysis.operating_expenses.taxes,
            analysis.operating_expenses.utilities,
            analysis.operating_expenses.other,
        ])
        
        for year in range(1, 11):
            row = year + 3
            
            yearly_rent = analysis.calculations.gross_rental_income * (1 + rent_growth) ** (year - 1)
            yearly_expenses = base_expenses * (1 + expense_growth) ** (year - 1)
            yearly_noi = yearly_rent - yearly_expenses
            yearly_debt_service = analysis.calculations.monthly_debt_service * 12
            yearly_cash_flow = yearly_noi - yearly_debt_service
            cumulative_cf += yearly_cash_flow
            
            data = [
                year,
                round(yearly_rent),
                round(yearly_expenses),
                round(yearly_noi),
                round(yearly_debt_service),
                round(yearly_cash_flow),
                round(cumulative_cf)
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col > 1:  # Format currency columns
                    cell.number_format = '"$"#,##0'
    
    def _create_ratios_metrics(self, property: Property, analysis: UnderwritingAnalysis):
        """Create ratios and metrics worksheet."""
        ws = self.workbook.create_sheet("Ratios & Metrics")
        
        # Title
        ws['A1'] = "INVESTMENT RATIOS & METRICS"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Profitability Ratios
        ws[f'A{row}'] = "Profitability Ratios"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        gross_rent_multiplier = analysis.purchase_price / analysis.calculations.gross_rental_income
        
        profitability = [
            ("Cap Rate:", f"{analysis.calculations.cap_rate:.2f}%"),
            ("Cash-on-Cash Return:", f"{analysis.calculations.cash_on_cash_return:.2f}%"),
            ("Gross Rent Multiplier:", f"{gross_rent_multiplier:.2f}"),
        ]
        
        for label, value in profitability:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Risk Ratios
        ws[f'A{row}'] = "Risk Ratios"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ltv_ratio = (analysis.loan_amount / analysis.purchase_price) * 100
        
        risk_ratios = [
            ("Debt Service Coverage Ratio:", f"{analysis.calculations.debt_service_coverage:.2f}"),
            ("Loan-to-Value Ratio:", f"{ltv_ratio:.2f}%"),
        ]
        
        for label, value in risk_ratios:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Efficiency Ratios
        ws[f'A{row}'] = "Efficiency Ratios"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        total_expenses = sum([
            analysis.operating_expenses.management,
            analysis.operating_expenses.maintenance,
            analysis.operating_expenses.insurance,
            analysis.operating_expenses.taxes,
            analysis.operating_expenses.utilities,
            analysis.operating_expenses.other,
        ])
        
        expense_ratio = (total_expenses / analysis.calculations.gross_rental_income) * 100
        noi_margin = (analysis.calculations.net_operating_income / analysis.calculations.gross_rental_income) * 100
        
        efficiency = [
            ("Operating Expense Ratio:", f"{expense_ratio:.2f}%"),
            ("NOI Margin:", f"{noi_margin:.2f}%"),
        ]
        
        for label, value in efficiency:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Per Unit Analysis
        ws[f'A{row}'] = "Per Unit Analysis"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        per_unit = [
            ("Price per Unit:", f"${analysis.purchase_price / property.units:,.0f}"),
            ("NOI per Unit:", f"${analysis.calculations.net_operating_income / property.units:,.0f}"),
            ("Cash Flow per Unit:", f"${analysis.calculations.annual_cash_flow / property.units:,.0f}"),
        ]
        
        for label, value in per_unit:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1